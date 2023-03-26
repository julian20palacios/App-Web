from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from .forms import TaskForm, VendorForm, ProductosForm, CompraOrdenForm, ProductosProveedorForm, NuevoModeloForm, PersonaForm
from .models import Task, Vendor, Productos, CompraOrden, ProductosProveedor, NuevoModelo, Persona
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime
from reportlab.pdfgen import canvas
import re
from decimal import Decimal
from django.db.models import Q, F
from django.http import JsonResponse
from django.contrib import messages
import pandas as pd
from django.db import transaction
from django.template.loader import get_template
from django.conf import settings
from django.templatetags.static import static
import matplotlib.pyplot as plt
from io import BytesIO
import base64, json




# ------------------------------- Graficar -------------------------------# 

def grafica(request):
    productos = ProductosProveedor.objects.all()

    
    nombres = [producto.proveedor for producto in productos]
    precios = [producto.precio for producto in productos]

    
    fig, ax = plt.subplots()
    ax.bar(nombres, precios)
    ax.set_xlabel('Nombre de Proveedor')
    ax.set_ylabel('Precios')
    ax.set_title('Análisis Gráfico')

    
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8').replace('\n', '')

    
    context = {
        'image': image_base64
    }

    # Renderizar la plantilla
    return render(request, 'grafica.html', context)

# HighChart
def chart(request):
    return render(request, 'chart.html')

def analisis_datos(request):
    # Obtener los datos de los productos del proveedor
    productos = ProductosProveedor.objects.all().values()
    # Convertir los datos a una lista de objetos para enviar al cliente
    productos_list = list(productos)
    # Devolver los datos como una respuesta JSON
    return JsonResponse({'productos': productos_list})




# ------------------------------- Inicio de la APP -------------------------------# 
def signup(request):
    if request.method == "GET":
        return render(request, "signup.html",{
            "formdecrearusuario": UserCreationForm
        })
    else:
        if request.POST["password1"] == request.POST["password2"]:
            username = request.POST["username"]
            if "@" in username and "." in username:
                password = request.POST["password1"]
                if sum(c.isdigit() for c in password) >= 5:
                    try:
                        user = User.objects.create_user(username=username, password=password)
                        user.save()
                        login(request, user)
                        return redirect('/tasks/')
                    except:
                        return render(request, "signup.html", {
                            "formdecrearusuario": UserCreationForm,
                            "error": 'El usuario ya existe'
                        })
                else:
                    return render(request, "signup.html", {
                        "formdecrearusuario": UserCreationForm,
                        "error": 'La contraseña debe tener al menos 5 números'
                    })
            else:
                return render(request, "signup.html", {
                    "formdecrearusuario": UserCreationForm,
                    "error": 'Correo incorrecto, debe ser un email válido'
                })
        else:
            return render(request, "signup.html",{
                "formdecrearusuario": UserCreationForm,
                "error": "Las contraseñas no coinciden"
            })
            
def home(request):   
    return render(request, "home.html")
 
def signin(request):
    if request.method == "GET":
        return render(request, "signin.html", {
            "autenticar": AuthenticationForm
        })
    else:
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is None:
            try:
                user = User.objects.get(username=username)
                if not user.check_password(password):
                    error = "Contraseña incorrecta, intenta de nuevo"
                else:
                    error = "Correo electrónico incorrecto"
            except User.DoesNotExist:
                error = "Usuario o contraseña incorrectos"
            return render(request, "signin.html", {
                "autenticar": AuthenticationForm,
                "error": error
            })
        else:
            login(request, user)
            return redirect("/aboutme/")
              
          
          
# ------------------------------- Mensajes de Confirmación -------------------------------# 
@login_required  
def aacongratulations1(request):   
    return render(request, "aacongratulations1.html")

@login_required  
def aacongratulations2(request):   
    return render(request, "aacongratulations2.html")




# ------------------------------- Tareas -------------------------------# 
@login_required
def tasks(request):
    tasks = Task.objects.filter(user=request.user)
    return render(request, "tasks.html", {"tasks" : tasks})

@login_required            
def create_task (request):   
    if request.method== "GET":
     return render(request, "create_task.html",{
       "form" :TaskForm        
      }) 
    else:
        try:  
          form = TaskForm(request.POST)
          new_task = form.save(commit=False)
          new_task.user = request.user
          new_task.save()
          return redirect("tareas")
        except ValueError:
         return render(request, "create_task.html",{
           "form" :TaskForm,
           "error" : "Por favor datos bien, si?"       
          }) 
 
@login_required
def task_detail(request, task_id):   
   if request.method == "GET":
        task = get_object_or_404(Task,pk=task_id, user=request.user)
        form = TaskForm(instance=task)
        return render(request, "task_detail.html",{"task":task, "form" : form}) 
   else:
        task = get_object_or_404(Task,pk=task_id, user=request.user)
        form = TaskForm(request.POST, instance=task)
        form.save()
        return redirect("/tasks")    
  
@login_required 
def task_delete(request, task_id):
    task = get_object_or_404(Task, pk=task_id, user=request.user)
    if request.method == "POST":
        task.delete()
        return redirect("/tasks")    

@login_required  
def complete_task(request, task_id):
    task = get_object_or_404(Task, pk=task_id, user=request.user)
    task.datecomplated = timezone.now()
    task.save()
    return redirect("/tasks")

@login_required
def eliminar_tarea(request, pk):
    tarea = get_object_or_404(Task, pk=pk)
    tarea.delete()
    return redirect('tareas') 

@login_required 
def export_to_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="tareas.xlsx"'

    # Crear el libro de Excel
    wb = Workbook()

    # Seleccionar la hoja de trabajo
    ws = wb.active

    # Escribir los encabezados
    ws['A1'] = 'Título'
    ws['B1'] = 'Descripción'
    ws['C1'] = 'Creado'
    ws['D1'] = 'Completado'
    ws['E1'] = 'Importante'
    ws['F1'] = 'Usuario'

    # Escribir los datos de las tareas
    tasks = Task.objects.all().values_list('title', 'description', 'created', 'datecomplated', 'important', 'user__username')
    for row_num, row in enumerate(tasks, start=2):
        for col_num, value in enumerate(row):
            # Convertir los objetos datetime a objetos datetime de Python sin zona horaria
            if isinstance(value, datetime):
                value = value.astimezone().replace(tzinfo=None)
            ws.cell(row=row_num, column=col_num+1, value=value)

    # Guardar el libro de Excel en la respuesta HTTP
    wb.save(response)
    return response 




# ------------------------------- Descripción y Salida -------------------------------# 
@login_required
def signout(request):
    logout(request)
    return redirect("/")

@login_required  
def about(request):   
    return render(request, "about.html") 
   



# ------------------------------- Proveedor or Vendor -------------------------------# 
@login_required            
def crear_vendor (request):   
    if request.method== "GET":
     return render(request, "crear_vendor.html",{
       "form" :VendorForm      
      }) 
    else:
        try:  
          form = VendorForm(request.POST)
          new_vendor = form.save(commit=False)
          new_vendor.user = request.user
          new_vendor.save()
          return redirect("/vendor")
        except ValueError:
         return render(request, "crear_vendor.html",{
           "form" :VendorForm,
           "error" : "Por favor datos bien, si?"       
          }) 

@login_required 
def detalle_proveedor(request, pk):
    proveedor = get_object_or_404(Vendor, pk=pk)
    return render(request, 'detalle_proveedor.html', {'vendor': proveedor})

@login_required
def vistavendor(request):
    vistavendor = Vendor.objects.all()
    return render(request, "vendor.html", {"vistavendor" : vistavendor})

@login_required
def eliminar_vendor(request, pk):
    vendor = get_object_or_404(Vendor, pk=pk)
    vendor.delete()
    return redirect('vendor')
  
@login_required
def export_vendors_to_excel(request):
    # Crear la respuesta HTTP
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="vendors.xlsx"'

    # Crear el libro de Excel
    wb = Workbook()

    # Seleccionar la hoja de trabajo
    ws = wb.active

    # Escribir los encabezados
    ws['A1'] = 'ID'
    ws['B1'] = 'Ciudad'
    ws['C1'] = 'País'
    ws['D1'] = 'Nombre de la empresa'
    ws['E1'] = 'ID de impuestos'
    ws['F1'] = 'Límite de crédito'
    ws['G1'] = 'Dirección'
    ws['H1'] = 'Correo electrónico'
    ws['I1'] = 'Nombre de contacto'
    ws['J1'] = 'Número de contacto'
    ws['K1'] = 'Sitio web'
    ws['L1'] = 'Método de entrega'
    ws['M1'] = 'Marca potencial'
    ws['N1'] = 'Categoría potencial'
    ws['O1'] = 'Oportunidad de precio'
    ws['P1'] = 'Productos destacados'
    ws['Q1'] = 'Verificación de crédito'
    ws['R1'] = 'Verificación de documentos'
    ws['S1'] = 'Verificación de factoring'
    ws['T1'] = 'Importante'
    ws['U1'] = 'Comprado'

    # Escribir los datos de los proveedores
    vendors = Vendor.objects.all().values_list(
        'id',
        'city',
        'country',
        'company_name',
        'tax_id',
        'credit_limit',
        'address',
        'email',
        'contact_name',
        'contact_number',
        'website',
        'delivery_method',
        'potential_brand',
        'potential_category',
        'price_opportunity',
        'featured_products',
        'credit_check',
        'document_check',
        'factoring_check',
        'is_important',
        'is_purchased',
    )
    for row_num, row in enumerate(vendors, start=2):
        for col_num, value in enumerate(row):
            # Convertir los objetos Decimal a flotantes de Python
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_num, column=col_num+1, value=value)

    # Guardar el libro de Excel en la respuesta HTTP
    wb.save(response)
    return response

@login_required
def get_proveedor_details(request):
    proveedor_id = request.GET.get('id')
    proveedor = Vendor.objects.get(id=proveedor_id)
    proveedor_details = {
        'correo_proveedor': proveedor.email,
        'direccion_proveedor': proveedor.address,
        'pais_proveedor': proveedor.country,
        'ciudad_proveedor': proveedor.city,
        'tax_id': proveedor.tax_id 
    }
    return JsonResponse(proveedor_details)
 
 
 
 

# ------------------------------- Ordenes de Compra -------------------------------# 
@login_required
def eliminar_orden_compra(request, pk):
    orden_compra = get_object_or_404(CompraOrden, pk=pk)
    orden_compra.delete()
    return redirect('ordenes')

@login_required 
def ordenes(request):
    compraorden = CompraOrden.objects.all()
    return render(request, 'ordenescompra.html', {'compraorden': compraorden})

@login_required 
def detalle_orden_compra(request, pk):
    orden_compra = get_object_or_404(CompraOrden, pk=pk)
    return render(request, 'detalle_oc.html', {'orden_compra': orden_compra})

@login_required
def crear_orden_compra(request):
    if request.method == 'POST':
        form = CompraOrdenForm(request.POST)
        if form.is_valid():
            compra_orden = form.save(commit=False)
            compra_orden.save()
            messages.success(request, 'La orden de compra se ha creado exitosamente.')
            return redirect('ordenes')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = CompraOrdenForm()
    return render(request, 'crear_orden_compra.html', {'form': form})

@login_required
def exportar_compraorden_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="compraorden.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Orden de compra"

    # Cabecera de la tabla
    ws['A1'] = 'Número de orden de compra'
    ws['B1'] = 'SKU'
    ws['C1'] = 'Producto'
    ws['D1'] = 'Precio del producto (USD)'
    ws['E1'] = 'Correo del proveedor'
    ws['F1'] = 'Dirección del proveedor'
    ws['G1'] = 'Tax ID'
    ws['H1'] = 'Ciudad del proveedor'
    ws['I1'] = 'Nombre del contacto del proveedor'
    ws['J1'] = 'Número del contacto del proveedor'
    ws['K1'] = 'Método de entrega del proveedor'
    ws['L1'] = 'UPC'
    ws['M1'] = 'Proveedor'
    


    # Obtener los datos de todas las CompraOrden
    compra_ordenes = CompraOrden.objects.all().values_list(
        'numero_orden_compra',
        'sku',
        'nombre_producto__nombre_producto',
        'precio_producto',
        'correo_proveedor',
        'direccion_proveedor',
        'tax_id',
        'ciudad_proveedor',
        'nombre_contacto_proveedor',
        'numero_contacto_proveedor',
        'metodo_entrega_proveedor',
        'upc',
        'proveedor',
          
    )

    # Agregar los datos a la tabla
    for row_num, row in enumerate(compra_ordenes, start=2):
        for col_num, value in enumerate(row):
            if isinstance(value, float):
                # Redondear los valores decimales a dos decimales
                value = round(value, 2)
            ws.cell(row=row_num, column=col_num+1, value=value)

    wb.save(response)
    return response




# ------------------------------- Productos Internos -------------------------------# 

@login_required 
def detalle_producto(request, pk):
    producto = get_object_or_404(Productos, pk=pk)
    return render(request, 'detalle_producto.html', {'producto': producto})

@login_required 
def productos(request):
    productos = Productos.objects.all()
    return render(request, 'producto.html', {'productos': productos})

@login_required            
def crear_producto (request):   
    if request.method== "GET":
     return render(request, "crear_productos.html",{
       "form" :ProductosForm      
      }) 
    else:
        try:  
          form = ProductosForm(request.POST)
          new_producto = form.save(commit=False)
          new_producto.user = request.user
          new_producto.save()
          return redirect("/productos")
        except ValueError:
         return render(request, "crear_productos.html",{
           "form" :ProductosForm,
           "error" : "Por favor datos bien, si?"       
          }) 
       
@login_required
def exportar_productos_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Nombre del producto'
    ws['B1'] = 'SKU'
    ws['C1'] = 'UPC'
    ws['D1'] = 'Precio (USD)'
    ws['E1'] = 'Categoría'
    ws['F1'] = 'Marca'
    ws['G1'] = 'Proveedor sugerido'
    ws['H1'] = 'Arancel'
    ws['I1'] = 'Importante'

    productos = Productos.objects.all().values_list(
        'nombre_producto',
        'sku',
        'upc',
        'precio_usd',
        'categoria',
        'marca',
        'proveedor_sugerido',
        'arancel',
        'check_importante',
    )
    for row_num, row in enumerate(productos, start=2):
        for col_num, value in enumerate(row):
            
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_num, column=col_num+1, value=value)
    wb.save(response)
    return response

@login_required
def eliminar_producto(request, pk):
    producto = get_object_or_404(Productos, pk=pk)
    producto.delete()
    return redirect('productos')





# ------------------------------- Publicaciones -------------------------------# 
@login_required  
def crear_persona(request):
    if request.method == 'POST':
        form = PersonaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('personas')
    else:
        form = PersonaForm()
    return render(request, 'crear_persona.html', {'form': form})

@login_required  
def personas(request):
    personas = Persona.objects.all()
    media_url = settings.MEDIA_URL
    return render(request, 'personas.html', {'personas': personas, 'media_url': media_url})





# -------------------------------Productos de Proveedor -------------------------------# 


@login_required 
def buscar_productos_proveedor(request):
    productoproveedores = ProductosProveedor.objects.all()
    nombre_query = request.GET.get('nombre')
    if nombre_query:
        productoproveedores = productoproveedores.filter(nombre__icontains=nombre_query)
    context = {
        'productoproveedores': productoproveedores,
        'nombre_query': nombre_query,
    }
    return render(request, 'productosproveedores.html', context)

@login_required
def product_vendor(request):
    productoproveedor = ProductosProveedor.objects.filter
    return render(request, "productsvendor.html", {"productoproveedor" : productoproveedor})

@login_required
def producto_vendor(request):
    productoproveedor = ProductosProveedor.objects.filter
    return render(request, "productosproveedores.html", {"productoproveedor" : productoproveedor})  
   
@login_required
def exportar_productos_vendor_excel(request):
    # Obtener todos los objetos ProductosProveedor
    productos = ProductosProveedor.objects.all()

    # Crear un libro de Excel y una hoja
    wb = Workbook()
    ws = wb.active

    # Agregar encabezados a la hoja
    ws['A1'] = 'Nombre'
    ws['B1'] = 'Proveedor'
    ws['C1'] = 'Precio'
    ws['D1'] = 'Cantidad Disponible'
    ws['E1'] = 'UPC'

    # Agregar los datos de los objetos a la hoja
    for index, producto in enumerate(productos):
        row_number = index + 2
        ws.cell(row=row_number, column=1, value=producto.nombre)
        ws.cell(row=row_number, column=2, value=producto.proveedor)
        ws.cell(row=row_number, column=3, value=producto.precio)
        ws.cell(row=row_number, column=4, value=producto.cantidad_disponible)
        ws.cell(row=row_number, column=5, value=producto.upc)

    # Crear una respuesta de HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="productos_proveedor.xlsx"'
    wb.save(response)

    return response

@login_required            
def crear_producto_proveedor (request):   
    if request.method== "GET":
     return render(request, "crearpodprove.html",{
       "form" :ProductosProveedorForm      
      }) 
    else:
        try:  
          form = ProductosProveedorForm(request.POST)
          new_producto_p = form.save(commit=False)
          new_producto_p.user = request.user
          new_producto_p.save()
          return redirect("product_vendor")
        except ValueError:
         return render(request, "crearpodprove.html",{
           "form" :ProductosProveedorForm,
           "error" : "Por favor datos bien, si?"       
          })     
   
@login_required
def importar_datos(request):
    if request.method == 'POST' and request.FILES['archivo']:
        archivo = request.FILES['archivo']
        df = pd.read_excel(archivo)
        for index, row in df.iterrows():
            try:
                upc = int(row['upc'])
            except ValueError:
                upc = 0 
            producto = ProductosProveedor(
                nombre=row['nombre'],
                proveedor=row['proveedor'],
                precio=row['precio'],
                cantidad_disponible=row['cantidad_disponible'],
                upc=upc
            )
            producto.save()
        return redirect('aacongratulations2')
    return render(request, 'productsvendor.html')     

@login_required
def eliminar_productos_general(request):
    if request.method == 'POST':
        ProductosProveedor.objects.all().delete()
    return redirect('product_vendor')



  
  
  
# --------------------------PRODUCTOS DEL MODELO -------------------------------# 
        
@login_required
def eliminar_datos(request):
    if request.method == 'POST':
        NuevoModelo.objects.all().delete()
        return redirect('productos/modelos') 
    return render(request, 'productos/modelos')

@login_required
def crear_nuevo_modelo(request):
    if request.method == 'POST':
        form = NuevoModeloForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('productos/modelos')  
    else:
        form = NuevoModeloForm()

    context = {
        'form': form,
    }
    return render(request, 'crear_nuevo_modelo.html', context)

@login_required
def nuevo_modelo_creado(request):
    nuevo_modelo = NuevoModelo.objects.last()
    context = {
        'nuevo_modelo': nuevo_modelo,
    }
    return render(request, 'aacongratulations1', context)

@login_required 
def productos_modelos(request):
    modelos = NuevoModelo.objects.all()
    return render(request, 'nuevo_modelo.html', {'modelos': modelos})

@login_required  
def actualizar_datos(request):
    # Obtener todos los objetos NuevoModelo
    objetos_nuevomodelo = NuevoModelo.objects.all()
    for objeto in objetos_nuevomodelo:
        # Buscar objeto ProductosProveedor con el precio más bajo para el nombre del objeto NuevoModelo
        producto_menor_precio = ProductosProveedor.objects.filter(nombre=objeto.nombre).order_by('precio').first()
        if producto_menor_precio:
            # Actualizar campos de NuevoModelo con los datos de ProductosProveedor
            objeto.proveedor = producto_menor_precio.proveedor
            objeto.precio = producto_menor_precio.precio
            objeto.cantidad_disponible = producto_menor_precio.cantidad_disponible
            objeto.upc = producto_menor_precio.upc
            objeto.save()
    # Renderizar la misma plantilla con un mensaje indicando que se actualizaron los datos
    return render(request, 'aacongratulations1.html', {'mensaje': 'Se actualizaron los datos.'})

@login_required 
def productos_modelo(request, nombre_modelo):
    productos = ProductosProveedor.objects.filter(nombre=nombre_modelo).order_by('precio')
    contexto = {'productos': productos}
    return render(request, 'filtroproductosvendor.html', contexto)

